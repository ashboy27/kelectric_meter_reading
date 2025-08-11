import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from datetime import datetime
import calendar

class ExcelExportService:
    def __init__(self):
        # Styling to match your app theme (Deep Teal #004D40, Soft Lilac #D8BFD8)
        self.header_fill = PatternFill(start_color="004D40", end_color="004D40", fill_type="solid")
        self.accent_fill = PatternFill(start_color="D8BFD8", end_color="D8BFD8", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True)
        self.title_font = Font(color="004D40", bold=True, size=14)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def create_yearly_excel(self, meter, yearly_data):
        """Create Excel file with summary + 12 monthly sheets"""
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Create summary sheet first
        self._create_summary_sheet(wb, meter, yearly_data)
        
        # Create monthly sheets (January 2025, February 2025, etc.)
        for month_data in yearly_data:
            self._create_monthly_sheet(wb, meter, month_data)
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer
    
    def _create_summary_sheet(self, wb, meter, yearly_data):
        """Create annual summary overview sheet"""
        ws = wb.create_sheet("Annual Summary", 0)
        year = yearly_data[0]['year'] if yearly_data else datetime.now().year
        
        # Header
        ws['A1'] = f"Annual Energy Report - {meter.name}"
        ws['A1'].font = Font(size=16, bold=True, color="004D40")
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Meter info
        ws['A3'] = "Meter Details"
        ws['A3'].font = self.title_font
        ws['A4'] = f"Meter ID: {meter.id}"
        ws['A5'] = f"Meter Name: {meter.name}"
        ws['A6'] = f"Report Year: {year}"
        ws['A7'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Monthly summary table
        ws['A9'] = "Monthly Overview"
        ws['A9'].font = self.title_font
        
        headers = ["Month", "Start Reading", "Total Entries", "Monthly Consumption", "Avg Daily"]
        for idx, header in enumerate(headers, 1):
            cell = ws.cell(11, idx, header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Fill monthly data
        total_consumption = 0
        total_entries = 0
        
        for idx, month_data in enumerate(yearly_data, 12):
            data = month_data['data']
            entries = data.get('entries', [])
            start_reading = data.get('start_reading', 0)
            
            # Calculate consumption
            end_reading = start_reading
            if entries:
                end_reading = max(entries, key=lambda x: x.reading_date).reading_value
            
            consumption = max(0, end_reading - start_reading)
            days_in_month = calendar.monthrange(month_data['year'], month_data['month'])[1]
            avg_daily = consumption / days_in_month if consumption > 0 else 0
            
            total_consumption += consumption
            total_entries += len(entries)
            
            ws.cell(idx, 1, month_data['month_name']).border = self.border
            ws.cell(idx, 2, round(start_reading, 2)).border = self.border
            ws.cell(idx, 3, len(entries)).border = self.border
            ws.cell(idx, 4, round(consumption, 2)).border = self.border
            ws.cell(idx, 5, round(avg_daily, 2)).border = self.border
        
        # Total row
        total_row = 24
        ws.cell(total_row, 1, "TOTAL").font = Font(bold=True)
        ws.cell(total_row, 1).fill = self.accent_fill
        ws.cell(total_row, 3, total_entries).font = Font(bold=True)
        ws.cell(total_row, 4, round(total_consumption, 2)).font = Font(bold=True)
        
        # Column widths
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 18
    
    def _create_monthly_sheet(self, wb, meter, month_data):
        """Create individual monthly sheet (e.g., 'January 2025')"""
        sheet_name = f"{month_data['month_name']} {month_data['year']}"
        ws = wb.create_sheet(sheet_name)
        
        data = month_data['data']
        entries = data.get('entries', [])
        start_reading = data.get('start_reading', 0)
        
        # Header
        ws['A1'] = f"{sheet_name} - Meter Readings"
        ws['A1'].font = Font(size=14, bold=True, color="004D40")
        ws.merge_cells('A1:E1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Month summary
        ws['A3'] = f"Meter: {meter.name}"
        ws['A4'] = f"Start Reading: {start_reading:.2f} units"
        ws['A5'] = f"Total Entries: {len(entries)}"
        
        if entries:
            end_reading = max(entries, key=lambda x: x.reading_date).reading_value
            consumption = max(0, end_reading - start_reading)
            ws['A6'] = f"End Reading: {end_reading:.2f} units"
            ws['A7'] = f"Total Consumption: {consumption:.2f} units"
        else:
            ws['A6'] = "No readings available for this month"
            return
        
        # Table headers
        ws['A9'] = "Daily Readings"
        ws['A9'].font = self.title_font
        
        headers = ["Date", "Time", "Reading Value", "Daily Change", "Notes"]
        for idx, header in enumerate(headers, 1):
            cell = ws.cell(11, idx, header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Sort entries by date
        sorted_entries = sorted(entries, key=lambda x: x.reading_date)
        prev_reading = start_reading
        
        # Fill data rows
        for idx, entry in enumerate(sorted_entries, 12):
            daily_change = entry.reading_value - prev_reading
            
            ws.cell(idx, 1, entry.reading_date.strftime("%Y-%m-%d")).border = self.border
            ws.cell(idx, 2, entry.reading_time.strftime("%H:%M:%S")).border = self.border
            ws.cell(idx, 3, round(entry.reading_value, 2)).border = self.border
            ws.cell(idx, 4, round(daily_change, 2)).border = self.border
            ws.cell(idx, 5, "Daily reading" if daily_change >= 0 else "Check reading").border = self.border
            
            prev_reading = entry.reading_value
        
        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20