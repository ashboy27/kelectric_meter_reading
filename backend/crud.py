# crud.py
from sqlmodel import Session, select
from fastapi import HTTPException
import datetime
import models, schemas
from database import engine
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
import calendar


from openpyxl.styles import PatternFill, Font, Alignment

# Define fill colors for thresholds (light to dark red)
fills = {
    170: PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),  # light red
    180: PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid"),  # medium red
    190: PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid"),  # dark red
    200: PatternFill(start_color="800000", end_color="800000", fill_type="solid"),  # very dark red
}

# Decide font color based on background (light bg => black text, dark bg => white text)
def get_font_color(change):
    if change >= 190:
        return Font(color="FFFFFF")  # white font for dark background
    else:
        return Font(color="000000")  # black font for light/medium background

# Determine which fill to apply
def get_fill(change):
    if change >= 200:
        return fills[200]
    elif change >= 190:
        return fills[190]
    elif change >= 180:
        return fills[180]
    elif change >= 170:
        return fills[170]
    else:
        return None  # no fill

def get_meters(household_token: str) -> list[schemas.MeterOut]:
    """
    Return a list of MeterOut for all meters in the given household,
    including computed `current_month_units`.
    """
    today = datetime.date.today()
    first_of_month = datetime.date(today.year, today.month, 1)

    with Session(engine) as sess:
        meters = sess.exec(
            select(models.Meter).where(models.Meter.household_token == household_token)
        ).all()

        result: list[schemas.MeterOut] = []
        for m in meters:
            # 1) total_units: most recent ever
            latest_value = (
                sess.exec(
                    select(models.Reading.reading_value)
                    .where(models.Reading.meter_id == m.id)
                    .order_by(models.Reading.reading_time.desc())
                    .limit(1)
                ).first()
                or 0.0
            )

            # 2) Determine start_val for this month
            sr = sess.exec(
                select(models.StartReading).where(
                    models.StartReading.meter_id == m.id,
                    models.StartReading.year == today.year,
                    models.StartReading.month == today.month,
                )
            ).one_or_none()

            if sr:
                start_val = sr.reading_value
            else:
                # Fallback to last reading before this month
                start_val = (
                    sess.exec(
                        select(models.Reading.reading_value)
                        .where(
                            models.Reading.meter_id == m.id,
                            models.Reading.reading_date < first_of_month,
                        )
                        .order_by(models.Reading.reading_time.desc())
                        .limit(1)
                    ).first()
                    or 0.0
                )

            # 3) Determine end_of_month: latest reading this month
            end_val = (
                sess.exec(
                    select(models.Reading.reading_value)
                    .where(
                        models.Reading.meter_id == m.id,
                        models.Reading.reading_date >= first_of_month,
                    )
                    .order_by(models.Reading.reading_time.desc())
                    .limit(1)
                ).first()
                or start_val
            )

            current_month_units = end_val - start_val

            result.append(
                schemas.MeterOut(
                    id=m.id,
                    name=m.name,
                    total_units=latest_value,
                    current_month_units=current_month_units,
                )
            )
        return result
def get_meter_by_id(meter_id: str):
    """Get meter by ID for Excel export"""
    with Session(engine) as sess:
        meter = sess.get(models.Meter, meter_id)
        return meter
    
def get_yearly_data_for_export(meter_id: str, year: int):
    """Get all 12 months of data for Excel export"""
    yearly_data = []
    
    for month in range(1, 13):
        try:
            # Use your existing get_monthly_data function
            month_data = get_monthly_data(meter_id, year, month)
            
            # Calculate additional metrics
            entries = month_data.entries
            start_reading = month_data.start_reading
            
            # Calculate consumption
            total_consumption = 0
            if entries:
                # Get the latest reading for the month
                latest_entry = max(entries, key=lambda x: x.time)
                total_consumption = latest_entry.reading - start_reading
            
            # Calculate average daily
            days_in_month = calendar.monthrange(year, month)[1]
            avg_daily = total_consumption / days_in_month if total_consumption > 0 else 0
            
            monthly_summary = {
                'month': month,
                'month_name': calendar.month_name[month],
                'year': year,
                'start_reading': start_reading,
                'entries': entries,
                'total_consumption': max(0, total_consumption),
                'average_daily': avg_daily,
                'total_readings': len(entries),
                'days_in_month': days_in_month,
                
            }
            
            yearly_data.append(monthly_summary)
            
        except Exception as e:
            print(f"Error getting data for {year}-{month}: {e}")
            # Handle months with no data
            yearly_data.append({
                'month': month,
                'month_name': calendar.month_name[month],
                'year': year,
                'start_reading': 0,
                'entries': [],
                'total_consumption': 0,
                'average_daily': 0,
                'total_readings': 0,
                'days_in_month': calendar.monthrange(year, month)[1]
            })
    
    return yearly_data

class ExcelExportService:
    def __init__(self):
        # Styling to match your app theme
        self.header_fill = PatternFill(start_color="004D40", end_color="004D40", fill_type="solid")
        self.accent_fill = PatternFill(start_color="D8BFD8", end_color="D8BFD8", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True)
        self.title_font = Font(color="004D40", bold=True, size=14)
        self.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
    
    def create_yearly_excel(self, meter, yearly_data):
        """Create Excel file with summary + 12 monthly sheets"""
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Create summary sheet first
        self._create_summary_sheet(wb, meter, yearly_data)
       
        # Create monthly sheets
        for month_data in yearly_data:
            self._create_monthly_sheet(wb, meter, month_data)
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        print("HRHEHEHREHEHEHRH")
        return excel_buffer
    
    def _create_summary_sheet(self, wb, meter, yearly_data):
        """Create annual summary sheet"""
        ws = wb.create_sheet("Annual Summary", 0)
        year = yearly_data[0]['year'] if yearly_data else datetime.datetime.now().year

        # Define fills for color coding (light to dark red)
        fills = {
            170: PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),  # light red
            180: PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid"),  # medium red
            190: PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid"),  # dark red
            200: PatternFill(start_color="800000", end_color="800000", fill_type="solid"),  # very dark red
        }

        def get_fill(value):
            if value >= 200:
                return fills[200]
            elif value >= 190:
                return fills[190]
            elif value >= 180:
                return fills[180]
            elif value >= 170:
                return fills[170]
            else:
                return None

        def get_font_color(value):
            # White font on dark backgrounds, else black
            if value >= 190:
                return Font(color="FFFFFF")
            else:
                return Font(color="000000")
        ws.column_dimensions['A'].width = 30

        # Header
        ws['A1'] = f"Annual Energy Report - {meter.name}"
        ws['A1'].font = Font(size=16, bold=True, color="004D40")
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal='center')

        # Meter info
        if meter.name == "First Floor Meter":
            meter.name = "FIRST FLOOR (SAY39286)"
        ws['A3'] = f"{meter.name}"
        ws['A4'] = f"Report Year: {year}"
        ws['A5'] = f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        # Monthly summary table headers
        headers = ["Month", "Start Reading", "Total Entries", "Monthly Consumption", "Avg Daily"]
        for idx, header in enumerate(headers, 1):
            cell = ws.cell(8, idx, header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border

        total_consumption = 0
        total_entries = 0
        latest_start_reading = 0

        # Fill monthly data rows
        for idx, month_data in enumerate(yearly_data, 9):
            entries = month_data.get('entries', [])
            start_reading = month_data.get('start_reading', 0)
            consumption = month_data.get('total_consumption', 0)

            # Track latest non-zero start reading
            if start_reading != 0:
                latest_start_reading = start_reading

            total_consumption += consumption
            total_entries += len(entries)

            ws.cell(idx, 1, month_data['month_name']).border = self.border

            start_cell = ws.cell(idx, 2, round(start_reading, 2))
            start_cell.border = self.border

            entries_cell = ws.cell(idx, 3, len(entries))
            entries_cell.border = self.border

            consumption_cell = ws.cell(idx, 4, round(consumption, 2))
            consumption_cell.border = self.border
            # Apply color coding for monthly consumption
            fill = get_fill(consumption)
            if fill:
                consumption_cell.fill = fill
                consumption_cell.font = get_font_color(consumption)

            avg_cell = ws.cell(idx, 5, round(month_data.get('average_daily', 0), 2))
            avg_cell.border = self.border

        # Total row
        total_row = 21
        ws.cell(total_row, 1, "TOTAL").font = Font(bold=True)
        ws.cell(total_row, 1).fill = self.accent_fill

        start_total_cell = ws.cell(total_row, 2, round(latest_start_reading, 2))
        start_total_cell.font = Font(bold=True)

        entries_total_cell = ws.cell(total_row, 3, total_entries)
        entries_total_cell.font = Font(bold=True)

        consumption_total_cell = ws.cell(total_row, 4, round(total_consumption, 0))
        consumption_total_cell.font = Font(bold=True)

        # Set column widths
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 20
        ws.column_dimensions['A'].width = 30

        # Add legend/explanation for color coding starting at row 23
        legend_start_row = 23
       

        ws.cell(legend_start_row + 1, 1, "≥ 200: Extreme consumption").fill = fills[200]
        ws.cell(legend_start_row + 2, 1, "≥ 190: Very high consumption").fill = fills[190]
        ws.cell(legend_start_row + 3, 1, "≥ 180: High consumption ").fill = fills[180]
        ws.cell(legend_start_row + 4, 1, "≥ 170: Elevated consumption ").fill = fills[170]

        # Adjust font color on legend cells for visibility
        for r in range(legend_start_row + 1, legend_start_row + 5):
            cell = ws.cell(r, 1)
            if cell.fill == fills[190] or cell.fill == fills[200]:
                cell.font = Font(color="FFFFFF")
            else:
                cell.font = Font(color="000000")
        

    def _create_monthly_sheet(self, wb, meter, month_data):
        """Create individual monthly sheet with prettier formatting"""
        try:
            print(f"DEBUG: Creating sheet for month_data: {month_data}")
            
            sheet_name = f"{month_data['month_name']} {month_data['year']}"
            print(f"DEBUG: Sheet name: {sheet_name}")
            
            ws = wb.create_sheet(sheet_name)
            
            entries = month_data.get('entries', [])
            start_reading = month_data.get('start_reading', 0)
            
            print(f"DEBUG: Found {len(entries)} entries, start_reading: {start_reading}")
            
            # Row 1: Main Header - METER READING SHEET {MONTH} {YEAR}
            ws['A1'] = f"METER READING BALANCE {month_data['month_name'].upper()} {month_data['year']}"
            ws['A1'].font = Font(size=16, bold=True, color="004D40")
            ws.merge_cells('A1:D1')
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Row 2: Floor and Meter Information using helper function
            if meter.name=="Second Floor Meter":
                floor_info="SECOND FLOOR (SCY74980)"
            else:
                floor_info = "FIRST FLOOR (SAY39286)"
            #floor_info = "First Floor Meter (SAY39286)" if meter.id!="fa76ead1-61a8-495d-8339-3abea2bf2740" else "Second Floor Meter (SCY74980)"
            ws['A2'] = floor_info
            print(floor_info)
            ws['A2'].font = Font(size=12, bold=True, color="666666")
            ws.merge_cells('A2:D2')
            ws['A2'].alignment = Alignment(horizontal='center')
            
            # Add some spacing and summary info below row 2
            ws['A4'] = f"Start Reading: {start_reading:.0f} units"
            ws['A4'].font = Font(size=10, bold=True,color="666666")
            ws['B4'] = f"Total Entries: {len(entries)}"
            ws['B4'].font = Font(size=10,bold=True, color="666666")
            ws['C4'] = f"Total Consumption: {month_data.get('total_consumption', 0):.0f} units"
            ws['C4'].font = Font(size=10,bold=True, color="666666")
            
            if not entries:
                print("DEBUG: No entries found for this month")
                ws['A6'] = "No readings available for this month"
                ws['A6'].font = Font(size=12, italic=True, color="999999")
                return
            
            # Print first entry for debugging
            print(f"DEBUG: First entry structure: {entries[0]}")
            print(f"DEBUG: First entry type: {type(entries[0])}")
            print(f"DEBUG: First entry attributes: {dir(entries[0])}")
            
            # Row 6: Table headers - starting from row 6 to give more space
            headers = ["Posted Timestamp", "Reading Value", "Units"]
            for idx, header in enumerate(headers, 1):
                cell = ws.cell(6, idx, header)
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = self.border
            
            # Sort entries by time (your entries have .time attribute)
            try:
                sorted_entries = sorted(entries, key=lambda x: x.time)
               
                print(f"DEBUG: Successfully sorted {len(sorted_entries)} entries")
            except Exception as sort_error:
                print(f"DEBUG: Error sorting entries: {sort_error}")
                sorted_entries = entries  # Use unsorted if sorting fails
            
            prev_reading = start_reading
            
            # Data rows start from row 7
            for idx, entry in enumerate(sorted_entries, 7):
                try:
                    print(f"DEBUG: Processing entry {idx-6}: {entry}")
                    
                    # Check if entry has required attributes
                    if not hasattr(entry, 'time'):
                        print(f"DEBUG: Entry missing 'time' attribute: {entry}")
                        continue
                    if not hasattr(entry, 'reading'):
                        print(f"DEBUG: Entry missing 'reading' attribute: {entry}")
                        continue
                        
                    change_from_prev = entry.reading - start_reading
                    print(f"DEBUG: Change calculated: {change_from_prev}")
                    
                    # Show the actual timestamp when reading was posted
                    try:
                        timestamp_str = entry.time.strftime("%Y-%m-%d %I:%M:%S %p")

                        print(f"DEBUG: Timestamp formatted: {timestamp_str}")
                    except Exception as time_error:
                        print(f"DEBUG: Error formatting time: {time_error}")
                        timestamp_str = str(entry.time)
                    
                    ws.cell(idx, 1, timestamp_str).border = self.border
                    ws.cell(idx, 2, round(entry.reading,0)).border = self.border

                    cell = ws.cell(idx, 3, round(change_from_prev, 0))
                    cell.border = self.border

                    fill = get_fill(change_from_prev)
                    if fill:
                        cell.fill = fill
                        cell.font = get_font_color(change_from_prev)
                    
                    
                    
                    
                    # Add alternating row colors for better readability
                    if idx % 2 == 0:  # Even rows
                        fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
                        for col in range(1, 3):
                            ws.cell(idx, col).fill = fill
                    
                    prev_reading = entry.reading
                    print(f"DEBUG: Successfully processed entry {idx-6}")
                    
                except Exception as entry_error:
                    print(f"DEBUG: Error processing entry {idx-6}: {entry_error}")
                    print(f"DEBUG: Problem entry: {entry}")
                    continue
            
            # Column widths - adjusted for prettier appearance
            ws.column_dimensions['A'].width = 30  # Timestamp column wider
            ws.column_dimensions['B'].width = 15  # Reading value
            ws.column_dimensions['C'].width = 18  # Change from previous

            
            # Add borders around the entire data area
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Apply border to header row
            for col in range(1, 4):
                ws.cell(6, col).border = thin_border
            
            print(f"DEBUG: Successfully completed sheet: {sheet_name}")
            
        except Exception as sheet_error:
            print(f"DEBUG: ERROR in _create_monthly_sheet: {sheet_error}")
            print(f"DEBUG: Sheet creation failed for month_data: {month_data}")
            print(f"DEBUG: Error type: {type(sheet_error)}")
            import traceback
            print(f"DEBUG: Full traceback: {traceback.format_exc()}")
            raise sheet_error  # Re-raise the caught exception

def create_excel_export(meter_id: str, year: int) -> BytesIO:
    """Main function to create Excel export - call this from your endpoint"""
    # Get meter info
    meter = get_meter_by_id(meter_id)
    if not meter:
        raise HTTPException(status_code=404, detail=f"Meter {meter_id} not found")
    
    # Get yearly data
    yearly_data = get_yearly_data_for_export(meter_id, year)
 
    # Generate Excel
    excel_service = ExcelExportService()
    excel_buffer = excel_service.create_yearly_excel(meter, yearly_data)
    
    return excel_buffer

def get_summary(household_token: str) -> schemas.HomeSummary:
    """
    Return aggregated summary for a household.
    """
    meters = get_meters(household_token)
    home_total = sum(m.total_units for m in meters)
    home_current = sum(m.current_month_units for m in meters)
    return schemas.HomeSummary(
        meters=meters, home_total=home_total, home_current_month=home_current
    )


def get_monthly_data(meter_id: str, year: int, month: int) -> schemas.MonthlyData:
    """
    Return the start reading and list of entries for a given meter/year/month.
    Each entry includes date, full timestamp, reading, and poster name.
    """
    with Session(engine) as sess:
        # 1) Fetch or default start reading
        sr = sess.exec(
            select(models.StartReading).where(
                models.StartReading.meter_id == meter_id,
                models.StartReading.year == year,
                models.StartReading.month == month,
            )
        ).one_or_none()
        start_val = sr.reading_value if sr else 0.0

        # 2) Date bounds
        first_day = datetime.date(year, month, 1)
        if month == 12:
            next_month = datetime.date(year + 1, 1, 1)
        else:
            next_month = datetime.date(year, month + 1, 1)

        # 3) Fetch readings in range
        entries = sess.exec(
            select(models.Reading)
            .where(
                models.Reading.meter_id == meter_id,
                models.Reading.reading_date >= first_day,
                models.Reading.reading_date < next_month,
            )
            .order_by(models.Reading.reading_time.desc())
        ).all()
       
        return schemas.MonthlyData(
            start_reading=start_val,
            entries=[
                schemas.EntryOut(
                    id=e.id,  
                    date=e.reading_date,
                    time=e.reading_time,
                    reading=e.reading_value,
                    posted_by=e.posted_by,
                )
                for e in entries
            ],
        )


def set_start_reading(meter_id: str, year: int, month: int, reading: float):
    """
    Explicitly set or reset the start reading for a meter/month.
    """
    with Session(engine) as sess:
        # Delete existing if any
        existing = sess.exec(
            select(models.StartReading).where(
                models.StartReading.meter_id == meter_id,
                models.StartReading.year == year,
                models.StartReading.month == month,
            )
        ).all()
        for sr in existing:
            sess.delete(sr)
        sess.commit()

        # Add new
        sr = models.StartReading(
            meter_id=meter_id, year=year, month=month, reading_value=reading
        )
        sess.add(sr)
        sess.commit()


def add_reading(
    meter_id: str,
    reading_date: datetime.date,
    reading_val: float,
    posted_by: str,
    reading_time: datetime.datetime
) -> int:
    
    reading_date = reading_date.replace(day=1)
    
    with Session(engine) as sess:
        m = sess.get(models.Meter, meter_id)
        if not m:
            raise HTTPException(status_code=404, detail="Meter not found")

        y, mo = reading_date.year, reading_date.month

        # Ensure start_reading exists
        sr = sess.exec(
            select(models.StartReading).where(
                models.StartReading.meter_id == meter_id,
                models.StartReading.year == y,
                models.StartReading.month == mo,
            )
        ).one_or_none()

        if not sr:
            raise HTTPException(
                status_code=400,
                detail="Cannot add entry: start reading not set for this month.",
            )

        # Ensure date matches selected month/year
        if y != sr.year or mo != sr.month:
            raise HTTPException(
                status_code=400,
                detail=f"Entry date {reading_date} must be within {sr.year}-{sr.month}.",
            )

        start_val = sr.reading_value

        # Get previous most entry for this month
        # prev_entry = sess.exec(
        #     select(models.Reading)
        #     .where(
        #         models.Reading.meter_id == meter_id,
        #         models.Reading.reading_date >= datetime.date(y, mo, 1),
        #         models.Reading.reading_date <= reading_date,
        #     )
        #     .order_by(
        #         models.Reading.reading_date.desc(), models.Reading.reading_time.desc()
        #     )
        # ).first()

        # prev_val = prev_entry.reading_value if prev_entry else start_val

        # # Enforce that new reading is not less than previous
        # if reading_val < prev_val:
        #     raise HTTPException(
        #         status_code=400,
        #         detail=f"New reading ({reading_val}) cannot be less than previous ({prev_val}).",
        #     )

        # Enforce freeze on primary meter
        new_total = reading_val - start_val

        level = 0
        if new_total > 200:
            level = 4
        elif new_total > 190:
            level = 3
        elif new_total > 180:
            level = 2
        elif new_total > 170:
            level = 1

        # Insert reading
        r = models.Reading(
            meter_id=m.id,
            reading_date=reading_date,
            reading_value=reading_val,
            posted_by=posted_by,
            reading_time=reading_time,
        )
        sess.add(r)
        sess.commit()
        return level


def has_start_reading(meter_id: str, year: int, month: int) -> bool:
    with Session(engine) as sess:
        exists = sess.exec(
            select(models.StartReading).where(
                models.StartReading.meter_id == meter_id,
                models.StartReading.year == year,
                models.StartReading.month == month,
            )
        ).first()
        return exists is not None

def delete_entry(entry_id: str):
    with Session(engine) as sess:
        entry = sess.get(models.Reading, entry_id)
        if not entry:
            raise HTTPException(status_code=404, detail="Entry not found")
        sess.delete(entry)
        sess.commit()


import datetime, pytz # type: ignore
from fastapi import FastAPI, Body # type: ignore

app = FastAPI()
PK_TZ = pytz.timezone("Asia/Karachi")
now_in_pk = datetime.datetime.now(PK_TZ)

@app.post("/entries")
def post_entry(
    posting_date: datetime.date = Body(...),
):
    """
    Posting date is YYYY-MM-DD format.
    If not provided, defaults to today in Pakistan timezone.
    We also add the current time in Pakistan timezone to the reading.
    """
    reading_time = datetime.datetime.combine(
        posting_date or now_in_pk.date(),
        now_in_pk.time(),
    )
    return {"reading_time": reading_time}

